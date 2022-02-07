#!/usr/bin/env python2.7
from jnpr.junos import Device
from jnpr.junos.exception import *
from argparse import ArgumentParser
from lxml import etree
from getpass import getpass
from openpyxl import Workbook
from openpyxl.styles import Font
from concurrent.futures import ThreadPoolExecutor, as_completed
import yaml
import logging
from datetime import datetime
from ncclient.operations.errors import TimeoutExpiredError
logging.basicConfig(
    level = logging.INFO
)

logging.getLogger("ncclient").setLevel(logging.WARNING)

def write_excel(Chassis):
        wb=Workbook()
        ws=wb.active
        ws.title='chassis_info'
        ws.cell(row=1, column=1).value = 'hostname'
        ws.cell(row=1, column=2).value = 'serial number'
        ws.cell(row=1, column=3).value = 'description'
        ws.cell(row=1, column=4).value = 'module name'
        ws.cell(row=1, column=5).value = 'module description'
        ws.cell(row=1, column=6).value = 'module serial number'
        ws.cell(row=1, column=7).value = 'module model'
        ws.cell(row=1, column=8).value = 'module clei-code'
        ws.cell(row=1, column=9).value = 'module version'
        ws.cell(row=1, column=10).value = 'module part number'
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).font = Font(bold=True)
        ws.cell(row=1, column=3).font = Font(bold=True)
        ws.cell(row=1, column=4).font = Font(bold=True)
        ws.cell(row=1, column=5).font = Font(bold=True)
        ws.cell(row=1, column=6).font = Font(bold=True)
        ws.cell(row=1, column=7).font = Font(bold=True)
        ws.cell(row=1, column=8).font = Font(bold=True)
        ws.cell(row=1, column=9).font = Font(bold=True)
        ws.cell(row=1, column=10).font = Font(bold=True)
        index=2
        index=2
        for chassis in Chassis.chassis_list:
                for sn in chassis.modules.keys():
                        ws.cell(row=index, column=1).value = chassis.device
                        ws.cell(row=index, column=2).value = chassis.serial_number
                        ws.cell(row=index, column=3).value = chassis.description
                        ws.cell(row=index, column=4).value = chassis.modules[sn]['name']
                        ws.cell(row=index, column=5).value = chassis.modules[sn]['description']
                        ws.cell(row=index, column=6).value = sn
                        ws.cell(row=index, column=7).value = chassis.modules[sn]['model']
                        ws.cell(row=index, column=8).value = chassis.modules[sn]['clei-code']
                        ws.cell(row=index, column=9).value = chassis.modules[sn]['version']
                        ws.cell(row=index, column=10).value = chassis.modules[sn]['part-number']
                        index = index+1
        wb.save('hardware_report.xlsx')

def connect(dev):

        try:
                with Device(user=login, host=dev[1], password=password, port=22) as device:
                        device.open(auto_probe=3, gather_facts=False)
                        device.timeout = 15

                        hostname = device.facts['hostname']
                        if hostname is None:
                                hostname = dev[0]
                        logging.info("{} thread conneced to {}, {}".format(datetime.now().time(), dev[0], dev[1]))
                        chassis_rpc = device.rpc.get_chassis_inventory()
                        name = chassis_rpc[0][0].text
                        serial_number = chassis_rpc[0][1].text
                        description = chassis_rpc[0][2].text
                        modules = {}
                        for entry in chassis_rpc.findall('.//chassis-module'):
                                if entry.findtext('serial-number') is not None:
                                        serial_number = entry.findtext('serial-number').replace('\n', '')
                                        modules[serial_number]={}
                                        modules[serial_number]['name'] = entry.findtext('name').replace('\n', '')
                                        modules[serial_number]['description'] = entry.findtext('description').replace('\n', '')
                                        if entry.findtext('model-number') is None:
                                                modules[serial_number]['model'] = 'BUILTIN'
                                        else:
                                                modules[serial_number]['model']  = entry.findtext('model-number').replace('\n', '')
                                        if entry.findtext('clei-code') is None:
                                                modules[serial_number]['clei-code'] = 'BUILTIN'
                                        else:
                                                modules[serial_number]['clei-code'] = entry.findtext('clei-code').replace('\n', '')
                                        if entry.findtext('version') is None:
                                                modules[serial_number]['version'] = 'BUILTIN'
                                        else:
                                                modules[serial_number]['version'] = entry.findtext('version').replace('\n', '')
                                        modules[serial_number]['part-number'] = entry.findtext('part-number').replace('\n', '')
                                        for sub in entry.findall('.//chassis-sub-module'):
                                                if sub.findtext('description') is not None:
                                                        modules[serial_number]['submodule-name']=sub.findtext('name').replace('\n', '')
                                                        modules[serial_number]['submodule-name']={}
                                                        modules[serial_number]['submodule-name']['description']=sub.findtext('description').replace('\n', '')
                                                        modules[serial_number]['submodule-name']['serial-number']=sub.findtext('serial-number').replace('\n', '')

                del device
                return({"hostname":hostname, "ip":dev[1], "sn":serial_number, "description":description, "modules":modules})



        except (TimeoutExpiredError, ConnectError, ProbeError, ConnectAuthError, ConnectTimeoutError, ConnectRefusedError, ConnectClosedError, RpcTimeoutError, IndexError) as err:
                logging.warning("{} connection failed to {}, {} due to {}".format(datetime.now().time(), dev[0], dev[1], err))
                return({dev[1]:"failed to connect to {} due to {}".format(dev[0], err)})

def main():

        global login, password
        parser = ArgumentParser()
        parser.add_argument('-l', '--LOGIN', type=str, help='login name to connect to device', required=True)
        #parser.add_argument('-d', '--DEVICE', type=str, help='ip address to connect to device', required=True)
        parser.add_argument('-f', '--FILE', type=str, help='csv file with devices ip addresses', required=True)
        args = parser.parse_args()
        login = args.LOGIN
        #ip = args.DEVICE
        dev_file = args.FILE
        password = getpass("Please enter your password:")
        devices = []


        with open(dev_file, 'r') as f:
                for a in f.readlines():
                        a = a.strip()
                        if len(a) > 1:
                                d, ip = a.split(',')
                                devices.append((d, ip))
                        else:
                                devices.append(('unknown', a[0]))

        with ThreadPoolExecutor(max_workers=10) as executor:
                results = executor.map(connect, devices)
                with open("chassis.yml", "w") as file:
                        for output in results:
                                yaml.dump(output, file)

        #write_excel(Chassis)

if __name__ == '__main__':
        main()
